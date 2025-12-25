"""
Script pour générer un plan de test Excel avec suivi OK/KO
"""

import pandas as pd
from datetime import datetime

# ============================================================================
# DÉFINITION DES TESTS
# ============================================================================

tests = [
    # ========== MODULE 1: AUTHENTICATION ==========
    {
        'ID': 'TEST-001',
        'Module': 'Authentication',
        'Scénario': 'Connexion utilisateur valide',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Utilisateur créé dans la base',
        'Étapes': '1. Aller sur /connexion/\n2. Saisir email + mot de passe\n3. Cliquer "Se connecter"',
        'Résultat attendu': 'Redirection vers /dashboard/',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-002',
        'Module': 'Authentication',
        'Scénario': 'Connexion - Mot de passe incorrect',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Utilisateur créé',
        'Étapes': '1. Aller sur /connexion/\n2. Saisir mauvais mot de passe\n3. Cliquer "Se connecter"',
        'Résultat attendu': 'Message erreur "Email ou mot de passe incorrect"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-003',
        'Module': 'Authentication',
        'Scénario': 'Déconnexion',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Utilisateur connecté',
        'Étapes': '1. Cliquer sur "Déconnexion"',
        'Résultat attendu': 'Redirection vers /connexion/',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-004',
        'Module': 'Authentication',
        'Scénario': 'Permissions par rôle',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Compte "lecteur" créé',
        'Étapes': '1. Se connecter comme lecteur\n2. Essayer de supprimer un élément',
        'Résultat attendu': 'Message "Accès refusé"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },

    # ========== MODULE 2: ENTREPRISES ==========
    {
        'ID': 'TEST-005',
        'Module': 'Entreprises',
        'Scénario': 'Créer une entreprise',
        'Priorité': 'HAUTE',
        'Pré-requis': 'Connecté comme Admin/Manager',
        'Étapes': '1. Aller sur /entreprises/\n2. Cliquer "Ajouter"\n3. Remplir le formulaire\n4. Enregistrer',
        'Résultat attendu': 'Message "✅ Entreprise ajoutée avec succès!"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-006',
        'Module': 'Entreprises',
        'Scénario': 'Modifier une entreprise',
        'Priorité': 'HAUTE',
        'Pré-requis': 'Entreprise existante',
        'Étapes': '1. Cliquer "Modifier"\n2. Changer le nom\n3. Enregistrer',
        'Résultat attendu': 'Message "✅ Entreprise mise à jour avec succès!"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-007',
        'Module': 'Entreprises',
        'Scénario': 'Supprimer une entreprise',
        'Priorité': 'MOYENNE',
        'Pré-requis': 'Connecté comme Admin uniquement',
        'Étapes': '1. Cliquer "Supprimer"\n2. Confirmer',
        'Résultat attendu': 'Message "🗑️ Entreprise supprimée avec succès!"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },

    # ========== MODULE 3: PERSONNEL ==========
    {
        'ID': 'TEST-008',
        'Module': 'Personnel',
        'Scénario': 'Ajouter un chauffeur',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Entreprise créée',
        'Étapes': '1. Aller sur /chauffeurs/\n2. Cliquer "Ajouter"\n3. Remplir: nom, prénom, téléphone, email, entreprise\n4. Enregistrer',
        'Résultat attendu': 'Message "✅ Chauffeur ajouté avec succès"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-009',
        'Module': 'Personnel',
        'Scénario': 'Affecter chauffeur à camion',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Chauffeur et camion disponibles',
        'Étapes': '1. Aller sur /affectations/\n2. Cliquer "Nouvelle affectation"\n3. Sélectionner chauffeur et camion\n4. Enregistrer',
        'Résultat attendu': 'Message "✅ Affectation créée avec succès!"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-010',
        'Module': 'Personnel',
        'Scénario': 'Terminer une affectation',
        'Priorité': 'HAUTE',
        'Pré-requis': 'Affectation active',
        'Étapes': '1. Cliquer "Terminer"\n2. Confirmer',
        'Résultat attendu': 'Camion redevient "disponible"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },

    # ========== MODULE 4: VÉHICULES ==========
    {
        'ID': 'TEST-011',
        'Module': 'Véhicules',
        'Scénario': 'Ajouter un camion',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Entreprise créée',
        'Étapes': '1. Aller sur /camions/\n2. Cliquer "Ajouter"\n3. Remplir: immatriculation, modèle, capacité, entreprise\n4. Enregistrer',
        'Résultat attendu': 'Message "✅ Camion ajouté avec succès!"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-012',
        'Module': 'Véhicules',
        'Scénario': 'Validation immatriculation unique',
        'Priorité': 'HAUTE',
        'Pré-requis': 'Camion avec immatriculation "ABC-123" existe',
        'Étapes': '1. Créer un camion avec la même immatriculation',
        'Résultat attendu': 'Message erreur "Cette immatriculation existe déjà"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-013',
        'Module': 'Véhicules',
        'Scénario': 'Ajouter un conteneur',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Client, transitaire, compagnie créés',
        'Étapes': '1. Aller sur /conteneurs/\n2. Cliquer "Ajouter"\n3. Remplir tous les champs\n4. Enregistrer',
        'Résultat attendu': 'Message "✅ Conteneur ajouté avec succès!"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },

    # ========== MODULE 5: COMMERCIAL ==========
    {
        'ID': 'TEST-014',
        'Module': 'Commercial',
        'Scénario': 'Créer un client',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Aucun',
        'Étapes': '1. Aller sur /clients/\n2. Cliquer "Ajouter"\n3. Remplir: nom, téléphone, email, adresse\n4. Enregistrer',
        'Résultat attendu': 'Client créé et visible dans la liste',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-015',
        'Module': 'Commercial',
        'Scénario': 'Créer un transitaire',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Aucun',
        'Étapes': '1. Aller sur /transitaires/\n2. Cliquer "Ajouter"\n3. Remplir les informations\n4. Enregistrer',
        'Résultat attendu': 'Transitaire créé',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },

    # ========== MODULE 6: CONTRATS & MISSIONS ==========
    {
        'ID': 'TEST-016',
        'Module': 'Contrats',
        'Scénario': 'Créer un contrat de transport',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Conteneur, client, transitaire, camion, chauffeur disponibles',
        'Étapes': '1. Aller sur /contrats/\n2. Cliquer "Créer"\n3. Remplir tous les champs (BL, montants, dates)\n4. Enregistrer',
        'Résultat attendu': 'Contrat créé, reliquat calculé, date limite +23j, PDF généré',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-017',
        'Module': 'Missions',
        'Scénario': 'Créer une mission',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Contrat créé',
        'Étapes': '1. Aller sur /missions/\n2. Cliquer "Créer"\n3. Sélectionner contrat\n4. Remplir destination, date\n5. Enregistrer',
        'Résultat attendu': 'Mission créée avec statut "en cours"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-018',
        'Module': 'Missions',
        'Scénario': 'Terminer une mission',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Mission "en cours"',
        'Étapes': '1. Cliquer "Terminer"\n2. Saisir date retour\n3. Confirmer',
        'Résultat attendu': 'Mission statut="terminée", Message succès, Log audit créé avec IP',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-019',
        'Module': 'Missions',
        'Scénario': 'Terminer mission en retard',
        'Priorité': 'HAUTE',
        'Pré-requis': 'Mission avec date retour > date limite',
        'Étapes': '1. Terminer avec date après limite\n2. Observer pénalité\n3. Confirmer',
        'Résultat attendu': 'Pénalité affichée (jours × 25,000 FCFA), Demande confirmation, Avertissement',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-020',
        'Module': 'Missions',
        'Scénario': 'Annuler une mission',
        'Priorité': 'HAUTE',
        'Pré-requis': 'Mission "en cours", Connecté Manager/Admin',
        'Étapes': '1. Cliquer "Annuler"\n2. Saisir raison\n3. Confirmer',
        'Résultat attendu': 'Mission annulée, Cautions annulées, Log audit créé',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },

    # ========== MODULE 7: FINANCES ==========
    {
        'ID': 'TEST-021',
        'Module': 'Finances',
        'Scénario': 'Créer une caution',
        'Priorité': 'HAUTE',
        'Pré-requis': 'Contrat existant',
        'Étapes': '1. Aller sur /cautions/\n2. Cliquer "Ajouter"\n3. Remplir contrat, montant, statut\n4. Enregistrer',
        'Résultat attendu': 'Caution créée',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-022',
        'Module': 'Finances',
        'Scénario': 'Créer un paiement mission',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Mission existante',
        'Étapes': '1. Aller sur /paiements/\n2. Cliquer "Créer"\n3. Remplir mission, montant, mode\n4. Enregistrer',
        'Résultat attendu': 'Paiement créé avec statut "en attente"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-023',
        'Module': 'Finances',
        'Scénario': 'Valider paiement (succès)',
        'Priorité': 'CRITIQUE',
        'Pré-requis': 'Mission terminée + Caution remboursée + Paiement en attente',
        'Étapes': '1. Cliquer "Valider"\n2. Confirmer',
        'Résultat attendu': 'Paiement validé, Message succès, Log audit avec IP, Notification envoyée',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-024',
        'Module': 'Finances',
        'Scénario': 'Valider paiement - Mission non terminée',
        'Priorité': 'HAUTE',
        'Pré-requis': 'Mission "en cours"',
        'Étapes': '1. Tenter de valider le paiement',
        'Résultat attendu': 'Message erreur "❌ Mission non terminée"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-025',
        'Module': 'Finances',
        'Scénario': 'Valider paiement - Caution non remboursée',
        'Priorité': 'HAUTE',
        'Pré-requis': 'Mission terminée + Caution bloquée',
        'Étapes': '1. Tenter de valider le paiement',
        'Résultat attendu': 'Message erreur "❌ Caution non remboursée"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },

    # ========== MODULE 8: RÉPARATIONS ==========
    {
        'ID': 'TEST-026',
        'Module': 'Réparations',
        'Scénario': 'Créer une réparation',
        'Priorité': 'MOYENNE',
        'Pré-requis': 'Camion et mécanicien disponibles',
        'Étapes': '1. Aller sur /reparations/\n2. Cliquer "Ajouter"\n3. Sélectionner camion, mécanicien(s)\n4. Remplir date, coût, description\n5. Enregistrer',
        'Résultat attendu': 'Réparation créée avec mécaniciens assignés',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-027',
        'Module': 'Réparations',
        'Scénario': 'Ajouter pièces à réparation',
        'Priorité': 'MOYENNE',
        'Pré-requis': 'Réparation existante',
        'Étapes': '1. Aller sur /pieces-reparees/\n2. Cliquer "Ajouter"\n3. Remplir nom, catégorie, quantité, coût, fournisseur\n4. Enregistrer',
        'Résultat attendu': 'Pièce ajoutée, Coût total = quantité × coût_unitaire',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },

    # ========== MODULE 9: SALAIRES ==========
    {
        'ID': 'TEST-028',
        'Module': 'Salaires',
        'Scénario': 'Créer fiche de salaire',
        'Priorité': 'HAUTE',
        'Pré-requis': 'Chauffeur existant, Connecté Admin/Manager/Comptable',
        'Étapes': '1. Aller sur /salaires/\n2. Cliquer "Créer"\n3. Sélectionner chauffeur, mois, année\n4. Remplir salaire base, primes, déductions\n5. Enregistrer',
        'Résultat attendu': 'Salaire net calculé automatiquement, Statut="brouillon"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-029',
        'Module': 'Salaires',
        'Scénario': 'Valider un salaire',
        'Priorité': 'HAUTE',
        'Pré-requis': 'Salaire en brouillon, Connecté Admin/Manager',
        'Étapes': '1. Cliquer "Valider"\n2. Confirmer',
        'Résultat attendu': 'Statut = "validé"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-030',
        'Module': 'Salaires',
        'Scénario': 'Marquer salaire comme payé',
        'Priorité': 'HAUTE',
        'Pré-requis': 'Salaire validé',
        'Étapes': '1. Cliquer "Marquer payé"\n2. Saisir mode paiement\n3. Confirmer',
        'Résultat attendu': 'Statut = "payé"',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },

    # ========== MODULE 10: DASHBOARD ==========
    {
        'ID': 'TEST-031',
        'Module': 'Dashboard',
        'Scénario': 'Affichage dashboard',
        'Priorité': 'HAUTE',
        'Pré-requis': 'Connecté',
        'Étapes': '1. Se connecter\n2. Observer dashboard',
        'Résultat attendu': 'Stats affichées, Graphiques visibles, Missions récentes listées',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-032',
        'Module': 'Dashboard',
        'Scénario': 'Filtrer stats par date',
        'Priorité': 'MOYENNE',
        'Pré-requis': 'Données existantes',
        'Étapes': '1. Sélectionner plage dates\n2. Cliquer "Filtrer"',
        'Résultat attendu': 'Statistiques recalculées pour la période',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-033',
        'Module': 'Dashboard',
        'Scénario': 'Stats avancées',
        'Priorité': 'MOYENNE',
        'Pré-requis': 'Connecté',
        'Étapes': '1. Aller sur /statistiques/',
        'Résultat attendu': 'Stats par camion, par chauffeur, évolution, top performers affichés',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },

    # ========== MODULE 11: AUDIT ==========
    {
        'ID': 'TEST-034',
        'Module': 'Audit',
        'Scénario': 'Consulter historique audit',
        'Priorité': 'HAUTE',
        'Pré-requis': 'Connecté Admin/Manager',
        'Étapes': '1. Aller sur /audit/',
        'Résultat attendu': 'Liste actions, Colonnes: Date, Utilisateur, Action, Modèle, Objet, IP',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-035',
        'Module': 'Audit',
        'Scénario': 'Filtrer logs audit',
        'Priorité': 'MOYENNE',
        'Pré-requis': 'Logs existants',
        'Étapes': '1. Filtrer par action\n2. Filtrer par utilisateur\n3. Filtrer par date',
        'Résultat attendu': 'Logs filtrés correctement',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-036',
        'Module': 'Notifications',
        'Scénario': 'Consulter notifications',
        'Priorité': 'MOYENNE',
        'Pré-requis': 'Connecté',
        'Étapes': '1. Aller sur /notifications/',
        'Résultat attendu': 'Notifications non lues en haut, Badge avec nombre',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-037',
        'Module': 'Notifications',
        'Scénario': 'Marquer toutes notifications lues',
        'Priorité': 'FAIBLE',
        'Pré-requis': 'Notifications non lues',
        'Étapes': '1. Cliquer "Tout marquer comme lu"',
        'Résultat attendu': 'Badge disparaît',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },

    # ========== MODULE 12: EXPORTS ==========
    {
        'ID': 'TEST-038',
        'Module': 'Exports',
        'Scénario': 'Export missions Excel',
        'Priorité': 'MOYENNE',
        'Pré-requis': 'Missions existantes',
        'Étapes': '1. Aller sur /missions/\n2. Cliquer "Exporter Excel"',
        'Résultat attendu': 'Fichier .xlsx téléchargé avec toutes les missions',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-039',
        'Module': 'Exports',
        'Scénario': 'Export missions PDF',
        'Priorité': 'MOYENNE',
        'Pré-requis': 'Missions existantes',
        'Étapes': '1. Cliquer "Exporter PDF"',
        'Résultat attendu': 'Fichier PDF téléchargé',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-040',
        'Module': 'Exports',
        'Scénario': 'Rapport paiements',
        'Priorité': 'MOYENNE',
        'Pré-requis': 'Paiements existants',
        'Étapes': '1. Aller sur /rapports/paiements/\n2. Sélectionner période\n3. Cliquer "Générer"',
        'Résultat attendu': 'Rapport PDF généré',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },

    # ========== MODULE 13: AUTO-SÉLECTION ==========
    {
        'ID': 'TEST-041',
        'Module': 'Auto-sélection',
        'Scénario': 'Sélection chauffeur → camion',
        'Priorité': 'MOYENNE',
        'Pré-requis': 'Affectation active',
        'Étapes': '1. Formulaire réparation\n2. Sélectionner un camion\n3. Observer champ chauffeur',
        'Résultat attendu': 'Chauffeur affecté sélectionné automatiquement',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
    {
        'ID': 'TEST-042',
        'Module': 'Auto-sélection',
        'Scénario': 'Sélection camion → chauffeur',
        'Priorité': 'MOYENNE',
        'Pré-requis': 'Affectation active',
        'Étapes': '1. Sélectionner un chauffeur\n2. Observer champ camion',
        'Résultat attendu': 'Camion affecté sélectionné automatiquement',
        'Résultat': '',
        'Commentaires': '',
        'Date test': '',
        'Testeur': '',
    },
]

# ============================================================================
# GÉNÉRATION DU FICHIER EXCEL
# ============================================================================

def generer_plan_test_excel():
    """Génère le plan de test Excel avec suivi OK/KO"""

    # Créer le DataFrame
    df = pd.DataFrame(tests)

    # Réorganiser les colonnes
    colonnes_ordre = [
        'ID', 'Module', 'Scénario', 'Priorité',
        'Pré-requis', 'Étapes', 'Résultat attendu',
        'Résultat', 'Commentaires', 'Date test', 'Testeur'
    ]
    df = df[colonnes_ordre]

    # Créer le fichier Excel
    output_file = 'PLAN_TEST_SUIVI.xlsx'

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Feuille principale
        df.to_excel(writer, sheet_name='Plan de test', index=False)

        # Récupérer le workbook et la feuille
        workbook = writer.book
        worksheet = writer.sheets['Plan de test']

        # Ajuster la largeur des colonnes
        worksheet.column_dimensions['A'].width = 12  # ID
        worksheet.column_dimensions['B'].width = 15  # Module
        worksheet.column_dimensions['C'].width = 35  # Scénario
        worksheet.column_dimensions['D'].width = 12  # Priorité
        worksheet.column_dimensions['E'].width = 30  # Pré-requis
        worksheet.column_dimensions['F'].width = 50  # Étapes
        worksheet.column_dimensions['G'].width = 50  # Résultat attendu
        worksheet.column_dimensions['H'].width = 10  # Résultat (OK/KO)
        worksheet.column_dimensions['I'].width = 40  # Commentaires
        worksheet.column_dimensions['J'].width = 12  # Date test
        worksheet.column_dimensions['K'].width = 20  # Testeur

        # Figer la première ligne
        worksheet.freeze_panes = 'A2'

        # Appliquer des styles
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Style pour l'en-tête
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Style pour les priorités
        priorite_colors = {
            'CRITIQUE': 'FF0000',  # Rouge
            'HAUTE': 'FFA500',     # Orange
            'MOYENNE': 'FFFF00',   # Jaune
            'FAIBLE': '90EE90'     # Vert clair
        }

        for row in range(2, len(tests) + 2):
            priorite = worksheet[f'D{row}'].value
            if priorite in priorite_colors:
                worksheet[f'D{row}'].fill = PatternFill(
                    start_color=priorite_colors[priorite],
                    end_color=priorite_colors[priorite],
                    fill_type='solid'
                )
                worksheet[f'D{row}'].font = Font(bold=True)

            # Aligner le texte
            for col in ['E', 'F', 'G', 'I']:
                worksheet[f'{col}{row}'].alignment = Alignment(
                    wrap_text=True,
                    vertical='top'
                )

        # Ajouter une feuille de statistiques
        stats_data = {
            'Priorité': ['CRITIQUE', 'HAUTE', 'MOYENNE', 'FAIBLE', 'TOTAL'],
            'Nombre': [
                len([t for t in tests if t['Priorité'] == 'CRITIQUE']),
                len([t for t in tests if t['Priorité'] == 'HAUTE']),
                len([t for t in tests if t['Priorité'] == 'MOYENNE']),
                len([t for t in tests if t['Priorité'] == 'FAIBLE']),
                len(tests)
            ],
            'Tests OK': ['', '', '', '', ''],
            'Tests KO': ['', '', '', '', ''],
            '% Réussite': ['', '', '', '', '']
        }

        df_stats = pd.DataFrame(stats_data)
        df_stats.to_excel(writer, sheet_name='Statistiques', index=False)

        # Ajuster la feuille statistiques
        ws_stats = writer.sheets['Statistiques']
        ws_stats.column_dimensions['A'].width = 15
        ws_stats.column_dimensions['B'].width = 12
        ws_stats.column_dimensions['C'].width = 12
        ws_stats.column_dimensions['D'].width = 12
        ws_stats.column_dimensions['E'].width = 15

        for cell in ws_stats[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    print(f"✅ Plan de test Excel généré avec succès: {output_file}")
    print(f"\n📊 Contenu:")
    print(f"  - Feuille 'Plan de test': {len(tests)} tests")
    print(f"  - Feuille 'Statistiques': Suivi de progression")
    print(f"\n📝 Colonnes:")
    print(f"  - Résultat: À remplir avec OK ou KO")
    print(f"  - Commentaires: Notes sur le test")
    print(f"  - Date test: Date d'exécution")
    print(f"  - Testeur: Votre nom")
    print(f"\n🎨 Priorités:")
    print(f"  🔴 CRITIQUE: {len([t for t in tests if t['Priorité'] == 'CRITIQUE'])} tests")
    print(f"  🟠 HAUTE: {len([t for t in tests if t['Priorité'] == 'HAUTE'])} tests")
    print(f"  🟡 MOYENNE: {len([t for t in tests if t['Priorité'] == 'MOYENNE'])} tests")
    print(f"  🟢 FAIBLE: {len([t for t in tests if t['Priorité'] == 'FAIBLE'])} tests")

if __name__ == '__main__':
    generer_plan_test_excel()
