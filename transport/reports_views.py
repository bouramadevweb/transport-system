"""
Vues pour les rapports financiers avancés
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Avg, Q
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime
from transport.models import (
    PaiementMission, Mission, Client, Chauffeur,
    Cautions, PrestationDeTransports
)
from transport.permissions import role_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils import get_column_letter


@login_required
@role_required('ADMIN', 'COMPTABLE')
def financial_reports(request):
    """
    Page principale des rapports financiers
    """
    # Récupérer la période sélectionnée
    period = request.GET.get('period', '30')  # Par défaut 30 jours

    # Calculer les dates
    now = timezone.now()
    if period == 'all':
        start_date = None
    else:
        days = int(period)
        start_date = now - timedelta(days=days)

    # Tous les paiements de l'entreprise (validés + en attente)
    paiements_base = PaiementMission.objects.filter(
        mission__contrat__entreprise=request.user.entreprise
    )
    if start_date:
        paiements_base = paiements_base.filter(date_paiement__gte=start_date)

    paiements_valides = paiements_base.filter(est_valide=True)
    paiements_attente = paiements_base.filter(est_valide=False)

    # === STATISTIQUES GLOBALES ===
    ca_valide   = paiements_valides.aggregate(Sum('montant_total'))['montant_total__sum'] or 0
    ca_attente  = paiements_attente.aggregate(Sum('montant_total'))['montant_total__sum'] or 0
    commissions = paiements_valides.aggregate(Sum('commission_transitaire'))['commission_transitaire__sum'] or 0
    nb_valides  = paiements_valides.count()
    nb_attente  = paiements_attente.count()
    nb_total    = paiements_base.count()
    montant_moyen = paiements_valides.aggregate(Avg('montant_total'))['montant_total__avg'] or 0

    stats = {
        'total_ca'        : ca_valide + ca_attente,
        'ca_valide'       : ca_valide,
        'ca_en_attente'   : ca_attente,
        'total_commissions': commissions,
        'ca_net'          : ca_valide - commissions,
        'nombre_paiements': nb_total,
        'nb_valides'      : nb_valides,
        'nb_attente'      : nb_attente,
        'montant_moyen'   : montant_moyen,
    }

    # === CA PAR CLIENT (tous paiements) ===
    ca_par_client = paiements_base.values(
        'mission__contrat__client__nom',
        'mission__contrat__client__type_client'
    ).annotate(
        total=Sum('montant_total'),
        total_valide=Sum('montant_total', filter=Q(est_valide=True)),
        count=Count('pk_paiement')
    ).order_by('-total')[:10]

    # === CA PAR CHAUFFEUR (tous paiements) ===
    ca_par_chauffeur = paiements_base.values(
        'mission__contrat__chauffeur__nom',
        'mission__contrat__chauffeur__prenom'
    ).annotate(
        total=Sum('montant_total'),
        total_valide=Sum('montant_total', filter=Q(est_valide=True)),
        count=Count('pk_paiement')
    ).order_by('-total')[:10]

    # === RÉPARTITION PAR MODE DE PAIEMENT ===
    mode_paiement = paiements_base.values('mode_paiement').annotate(
        total=Sum('montant_total'),
        count=Count('pk_paiement')
    ).order_by('-total')

    # === ÉVOLUTION MENSUELLE (6 derniers mois, tous paiements) ===
    monthly_data = []
    for i in range(5, -1, -1):
        month_start = now - timedelta(days=30 * i)
        month_end   = now - timedelta(days=30 * (i - 1)) if i > 0 else now

        month_qs = PaiementMission.objects.filter(
            mission__contrat__entreprise=request.user.entreprise,
            date_paiement__gte=month_start,
            date_paiement__lt=month_end
        )
        ca_m        = month_qs.aggregate(Sum('montant_total'))['montant_total__sum'] or 0
        ca_valide_m = month_qs.filter(est_valide=True).aggregate(Sum('montant_total'))['montant_total__sum'] or 0
        monthly_data.append({
            'label'     : month_start.strftime('%b %Y'),
            'ca'        : ca_m,
            'ca_valide' : ca_valide_m,
            'ca_attente': ca_m - ca_valide_m,
            'count'     : month_qs.count()
        })

    # === MISSIONS EN ATTENTE DE PAIEMENT ===
    missions_en_attente = Mission.objects.filter(
        statut='terminée',
        paiementmission__isnull=True,
        contrat__entreprise=request.user.entreprise
    ).select_related('contrat__client', 'contrat__chauffeur').count()

    # === CAUTIONS EN COURS ===
    cautions_en_cours = Cautions.objects.filter(
        contrat__entreprise=request.user.entreprise,
        statut='en_attente'
    ).aggregate(total=Sum('montant'))['total'] or 0

    context = {
        'stats': stats,
        'ca_par_client': ca_par_client,
        'ca_par_chauffeur': ca_par_chauffeur,
        'mode_paiement': mode_paiement,
        'monthly_data': monthly_data,
        'missions_en_attente': missions_en_attente,
        'cautions_en_cours': cautions_en_cours,
        'period': period,
    }

    return render(request, 'transport/reports/financial.html', context)


@login_required
@role_required('ADMIN', 'COMPTABLE')
def export_financial_report_excel(request):
    """
    Export du rapport financier en Excel
    """
    # Récupérer les paramètres
    period = request.GET.get('period', '30')

    # Calculer les dates
    now = timezone.now()
    if period == 'all':
        start_date = None
        period_label = "Toutes périodes"
    else:
        days = int(period)
        start_date = now - timedelta(days=days)
        period_label = f"{days} derniers jours"

    # Filtrer les paiements
    paiements = PaiementMission.objects.filter(
        est_valide=True,
        mission__contrat__entreprise=request.user.entreprise
    ).select_related(
        'mission__contrat__client',
        'mission__contrat__chauffeur',
        'mission'
    )
    if start_date:
        paiements = paiements.filter(date_validation__gte=start_date)

    # Créer le fichier Excel
    wb = openpyxl.Workbook()

    # === FEUILLE 1: RÉSUMÉ ===
    ws_resume = wb.active
    ws_resume.title = "Résumé"

    # Titre
    ws_resume['A1'] = "RAPPORT FINANCIER"
    ws_resume['A1'].font = Font(size=18, bold=True, color="FFFFFF")
    ws_resume['A1'].fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    ws_resume.merge_cells('A1:D1')
    ws_resume['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_resume.row_dimensions[1].height = 30

    # Période
    ws_resume['A3'] = "Période:"
    ws_resume['B3'] = period_label
    ws_resume['A4'] = "Date de génération:"
    ws_resume['B4'] = now.strftime('%d/%m/%Y %H:%M')

    # Statistiques
    ws_resume['A6'] = "STATISTIQUES GLOBALES"
    ws_resume['A6'].font = Font(size=14, bold=True)

    stats_data = [
        ['Indicateur', 'Valeur'],
        ['CA Total', f"{paiements.aggregate(Sum('montant_total'))['montant_total__sum'] or 0:,.0f} FCFA"],
        ['Commissions', f"{paiements.aggregate(Sum('commission_transitaire'))['commission_transitaire__sum'] or 0:,.0f} FCFA"],
        ['CA Net', f"{(paiements.aggregate(Sum('montant_total'))['montant_total__sum'] or 0) - (paiements.aggregate(Sum('commission_transitaire'))['commission_transitaire__sum'] or 0):,.0f} FCFA"],
        ['Nombre de paiements', paiements.count()],
        ['Montant moyen', f"{paiements.aggregate(Avg('montant_total'))['montant_total__avg'] or 0:,.0f} FCFA"],
    ]

    for row_idx, row_data in enumerate(stats_data, start=8):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_resume.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 8:  # En-tête
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Ajuster les largeurs de colonnes
    ws_resume.column_dimensions['A'].width = 25
    ws_resume.column_dimensions['B'].width = 20

    # === FEUILLE 2: DÉTAILS DES PAIEMENTS ===
    ws_paiements = wb.create_sheet("Détails Paiements")

    # En-têtes
    headers = ['Date', 'N° Mission', 'Client', 'Chauffeur', 'Origine', 'Destination', 'Montant Total', 'Commission', 'Net', 'Mode Paiement']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_paiements.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Données
    for row_idx, paiement in enumerate(paiements.order_by('-date_validation'), start=2):
        client_nom = paiement.mission.contrat.client.nom if paiement.mission and paiement.mission.contrat and paiement.mission.contrat.client else "N/A"
        chauffeur_nom = f"{paiement.mission.contrat.chauffeur.nom} {paiement.mission.contrat.chauffeur.prenom}" if paiement.mission and paiement.mission.contrat and paiement.mission.contrat.chauffeur else "N/A"

        row_data = [
            paiement.date_validation.strftime('%d/%m/%Y') if paiement.date_validation else 'N/A',
            str(paiement.mission.pk_mission)[:20] if paiement.mission else 'N/A',
            client_nom,
            chauffeur_nom,
            paiement.mission.origine if paiement.mission else 'N/A',
            paiement.mission.destination if paiement.mission else 'N/A',
            float(paiement.montant_total) if paiement.montant_total else 0,
            float(paiement.commission_transitaire) if paiement.commission_transitaire else 0,
            float(paiement.montant_total or 0) - float(paiement.commission_transitaire or 0),
            paiement.mode_paiement.upper() if paiement.mode_paiement else 'N/A',
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_paiements.cell(row=row_idx, column=col_idx, value=value)
            # Format nombres
            if col_idx in [7, 8, 9]:  # Colonnes montants
                cell.number_format = '#,##0 "FCFA"'

    # Ajuster largeurs
    for col in range(1, len(headers) + 1):
        ws_paiements.column_dimensions[get_column_letter(col)].width = 15

    # === FEUILLE 3: CA PAR CLIENT ===
    ws_clients = wb.create_sheet("CA par Client")

    # En-têtes
    ws_clients['A1'] = "Client"
    ws_clients['B1'] = "Type"
    ws_clients['C1'] = "CA Total"
    ws_clients['D1'] = "Nombre"

    for cell in [ws_clients['A1'], ws_clients['B1'], ws_clients['C1'], ws_clients['D1']]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="764BA2", end_color="764BA2", fill_type="solid")

    # Données
    ca_clients = paiements.values(
        'mission__contrat__client__nom',
        'mission__contrat__client__type_client'
    ).annotate(
        total=Sum('montant_total'),
        count=Count('pk_paiement')
    ).order_by('-total')

    for row_idx, data in enumerate(ca_clients, start=2):
        ws_clients.cell(row=row_idx, column=1, value=data['mission__contrat__client__nom'] or 'N/A')
        ws_clients.cell(row=row_idx, column=2, value=data['mission__contrat__client__type_client'] or 'N/A')
        cell_total = ws_clients.cell(row=row_idx, column=3, value=float(data['total']) if data['total'] else 0)
        cell_total.number_format = '#,##0 "FCFA"'
        ws_clients.cell(row=row_idx, column=4, value=data['count'])

    # Largeurs
    ws_clients.column_dimensions['A'].width = 30
    ws_clients.column_dimensions['B'].width = 15
    ws_clients.column_dimensions['C'].width = 20
    ws_clients.column_dimensions['D'].width = 10

    # === RETOURNER LE FICHIER ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=rapport_financier_{now.strftime("%Y%m%d")}.xlsx'

    wb.save(response)
    return response


@login_required
@role_required('ADMIN', 'COMPTABLE')
def client_report(request, client_id):
    """
    Rapport détaillé pour un client spécifique
    """
    client = Client.objects.get(pk_client=client_id)

    # Missions du client
    missions = Mission.objects.filter(
        contrat__client=client
    ).select_related('contrat').order_by('-date_depart')

    # Paiements du client
    paiements = PaiementMission.objects.filter(
        mission__contrat__client=client,
        est_valide=True
    )

    # Statistiques
    stats = {
        'total_missions': missions.count(),
        'missions_terminee': missions.filter(statut='terminée').count(),
        'missions_en_cours': missions.filter(statut='en cours').count(),
        'total_ca': paiements.aggregate(Sum('montant_total'))['montant_total__sum'] or 0,
        'ca_moyen': paiements.aggregate(Avg('montant_total'))['montant_total__avg'] or 0,
    }

    context = {
        'client': client,
        'missions': missions[:20],  # 20 dernières missions
        'paiements': paiements[:10],  # 10 derniers paiements
        'stats': stats,
    }

    return render(request, 'transport/reports/client_detail.html', context)
