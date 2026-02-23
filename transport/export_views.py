"""
Vues pour l'export de données en Excel et CSV
==============================================
"""

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from datetime import datetime
from .models import Mission, PaiementMission


@login_required
def export_missions_excel(request):
    """Export des missions en format Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("❌ Bibliothèque openpyxl non installée. Exécutez: pip install openpyxl", status=500)

    from .filters import MissionFilter

    # Récupérer les missions filtrées
    missions = Mission.objects.select_related(
        'contrat', 'contrat__chauffeur', 'contrat__client', 'contrat__camion', 'contrat__conteneur'
    ).filter(contrat__entreprise=request.user.entreprise).order_by('-date_depart')

    # Appliquer les mêmes filtres que la liste
    missions = MissionFilter.apply(missions, request)

    # Créer un nouveau workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Missions"

    # Style de l'en-tête
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # En-têtes
    headers = [
        "ID Mission", "Statut", "Origine", "Destination",
        "Date Départ", "Date Retour", "Chauffeur", "Client",
        "Camion", "Conteneur"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Données
    for row, mission in enumerate(missions, 2):
        ws.cell(row=row, column=1).value = mission.pk_mission[:15]
        ws.cell(row=row, column=2).value = mission.statut
        ws.cell(row=row, column=3).value = mission.origine
        ws.cell(row=row, column=4).value = mission.destination
        ws.cell(row=row, column=5).value = mission.date_depart.strftime('%d/%m/%Y') if mission.date_depart else ''
        ws.cell(row=row, column=6).value = mission.date_retour.strftime('%d/%m/%Y') if mission.date_retour else ''

        if mission.contrat:
            ws.cell(row=row, column=7).value = f"{mission.contrat.chauffeur.nom} {mission.contrat.chauffeur.prenom}" if mission.contrat.chauffeur else ''
            ws.cell(row=row, column=8).value = mission.contrat.client.nom if mission.contrat.client else ''
            ws.cell(row=row, column=9).value = mission.contrat.camion.immatriculation if mission.contrat.camion else ''
            ws.cell(row=row, column=10).value = mission.contrat.conteneur.numero_conteneur if mission.contrat.conteneur else ''

    # Ajuster les largeurs de colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=missions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response


@login_required
def export_missions_csv(request):
    """Export des missions en format CSV"""
    import csv
    from .filters import MissionFilter

    # Récupérer les missions filtrées
    missions = Mission.objects.select_related(
        'contrat', 'contrat__chauffeur', 'contrat__client', 'contrat__camion', 'contrat__conteneur'
    ).filter(contrat__entreprise=request.user.entreprise).order_by('-date_depart')

    # Appliquer les mêmes filtres que la liste
    missions = MissionFilter.apply(missions, request)

    # Créer la réponse HTTP
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename=missions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

    # BOM UTF-8 pour Excel
    response.write('\ufeff')

    writer = csv.writer(response, delimiter=';')

    # En-têtes
    writer.writerow([
        'ID Mission', 'Statut', 'Origine', 'Destination',
        'Date Départ', 'Date Retour', 'Chauffeur', 'Client',
        'Camion', 'Conteneur'
    ])

    # Données
    for mission in missions:
        writer.writerow([
            mission.pk_mission[:15],
            mission.statut,
            mission.origine,
            mission.destination,
            mission.date_depart.strftime('%d/%m/%Y') if mission.date_depart else '',
            mission.date_retour.strftime('%d/%m/%Y') if mission.date_retour else '',
            f"{mission.contrat.chauffeur.nom} {mission.contrat.chauffeur.prenom}" if mission.contrat and mission.contrat.chauffeur else '',
            mission.contrat.client.nom if mission.contrat and mission.contrat.client else '',
            mission.contrat.camion.immatriculation if mission.contrat and mission.contrat.camion else '',
            mission.contrat.conteneur.numero_conteneur if mission.contrat and mission.contrat.conteneur else '',
        ])

    return response


@login_required
def export_paiements_excel(request):
    """Export des paiements en format Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("❌ Bibliothèque openpyxl non installée. Exécutez: pip install openpyxl", status=500)

    from .filters import PaiementMissionFilter

    # Récupérer les paiements filtrés
    paiements = PaiementMission.objects.select_related(
        'mission', 'mission__contrat__chauffeur', 'caution'
    ).filter(mission__contrat__entreprise=request.user.entreprise).order_by('-date_paiement')

    # Appliquer les mêmes filtres que la liste
    paiements = PaiementMissionFilter.apply(paiements, request)

    # Créer un nouveau workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paiements"

    # Style de l'en-tête
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # En-têtes
    headers = [
        "ID Paiement", "Mission", "Montant Total", "Commission",
        "Caution Retirée", "Est Validé", "Date Validation",
        "Chauffeur", "Mode Paiement"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Données
    for row, paiement in enumerate(paiements, 2):
        ws.cell(row=row, column=1).value = paiement.pk_paiement[:15]
        ws.cell(row=row, column=2).value = paiement.mission.pk_mission[:15] if paiement.mission else ''
        ws.cell(row=row, column=3).value = float(paiement.montant_total)
        ws.cell(row=row, column=4).value = float(paiement.commission_transitaire)
        ws.cell(row=row, column=5).value = "Oui" if paiement.caution_est_retiree else "Non"
        ws.cell(row=row, column=6).value = "Validé" if paiement.est_valide else "En attente"
        ws.cell(row=row, column=7).value = paiement.date_validation.strftime('%d/%m/%Y') if paiement.date_validation else ''

        if paiement.mission and paiement.mission.contrat and paiement.mission.contrat.chauffeur:
            chauffeur = paiement.mission.contrat.chauffeur
            ws.cell(row=row, column=8).value = f"{chauffeur.nom} {chauffeur.prenom}"

        ws.cell(row=row, column=9).value = paiement.mode_paiement

    # Ajuster les largeurs de colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=paiements_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    wb.save(response)
    return response


@login_required
def export_paiements_csv(request):
    """Export des paiements en format CSV"""
    import csv
    from .filters import PaiementMissionFilter

    # Récupérer les paiements filtrés
    paiements = PaiementMission.objects.select_related(
        'mission', 'mission__contrat__chauffeur', 'caution'
    ).filter(mission__contrat__entreprise=request.user.entreprise).order_by('-date_paiement')

    # Appliquer les mêmes filtres que la liste
    paiements = PaiementMissionFilter.apply(paiements, request)

    # Créer la réponse HTTP
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename=paiements_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

    # BOM UTF-8 pour Excel
    response.write('\ufeff')

    writer = csv.writer(response, delimiter=';')

    # En-têtes
    writer.writerow([
        'ID Paiement', 'Mission', 'Montant Total', 'Commission',
        'Caution Retirée', 'Est Validé', 'Date Validation',
        'Chauffeur', 'Mode Paiement'
    ])

    # Données
    for paiement in paiements:
        chauffeur = ''
        if paiement.mission and paiement.mission.contrat and paiement.mission.contrat.chauffeur:
            c = paiement.mission.contrat.chauffeur
            chauffeur = f"{c.nom} {c.prenom}"

        writer.writerow([
            paiement.pk_paiement[:15],
            paiement.mission.pk_mission[:15] if paiement.mission else '',
            float(paiement.montant_total),
            float(paiement.commission_transitaire),
            'Oui' if paiement.caution_est_retiree else 'Non',
            'Validé' if paiement.est_valide else 'En attente',
            paiement.date_validation.strftime('%d/%m/%Y') if paiement.date_validation else '',
            chauffeur,
            paiement.mode_paiement,
        ])

    return response


@login_required
def export_utilisateurs_excel(request):
    """Export des utilisateurs en format Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("❌ Bibliothèque openpyxl non installée. Exécutez: pip install openpyxl", status=500)

    from .models import Utilisateur
    from .permissions import get_user_role, ROLES

    # Récupérer tous les utilisateurs
    utilisateurs = Utilisateur.objects.exclude(
        pk_utilisateur=''
    ).exclude(
        pk_utilisateur__isnull=True
    ).filter(entreprise=request.user.entreprise).prefetch_related('groups').order_by('nom_utilisateur', 'email')

    # Créer un nouveau workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Utilisateurs"

    # Style de l'en-tête
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # En-têtes
    headers = [
        "Nom d'utilisateur", "Email", "Rôle", "Statut",
        "Staff", "Superuser", "Date de création", "Entreprise"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Données
    for row_idx, user in enumerate(utilisateurs, 2):
        # Déterminer le rôle
        role_code = get_user_role(user)
        if role_code == 'SUPERUSER':
            role_display = 'Super Admin'
        elif role_code in ROLES:
            role_display = ROLES[role_code]['name']
        else:
            role_display = 'Aucun rôle'

        # Nom d'affichage
        display_name = user.nom_utilisateur or user.email or f"User {user.pk_utilisateur}"

        ws.append([
            display_name,
            user.email,
            role_display,
            'Actif' if user.is_active else 'Inactif',
            'Oui' if user.is_staff else 'Non',
            'Oui' if user.is_superuser else 'Non',
            user.date_creation.strftime('%d/%m/%Y %H:%M') if user.date_creation else '',
            user.entreprise.nom if user.entreprise else '',
        ])

    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Créer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"utilisateurs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Sauvegarder le workbook dans la réponse
    wb.save(response)

    return response


@login_required
def export_audit_excel(request):
    """Export de l'historique d'audit en format Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("❌ Bibliothèque openpyxl non installée. Exécutez: pip install openpyxl", status=500)

    from .models import AuditLog

    # Récupérer les logs avec les mêmes filtres que la liste
    logs = AuditLog.objects.select_related('utilisateur').order_by('-timestamp')

    # Appliquer les filtres depuis les paramètres GET
    action_filter = request.GET.get('action')
    if action_filter and action_filter != 'tous':
        logs = logs.filter(action=action_filter)

    model_filter = request.GET.get('model')
    if model_filter:
        logs = logs.filter(model_name__icontains=model_filter)

    user_filter = request.GET.get('utilisateur')
    if user_filter:
        logs = logs.filter(utilisateur__pk_utilisateur=user_filter)

    date_debut = request.GET.get('date_debut')
    if date_debut:
        try:
            from django.utils.dateparse import parse_date
            logs = logs.filter(timestamp__date__gte=parse_date(date_debut))
        except:
            pass

    date_fin = request.GET.get('date_fin')
    if date_fin:
        try:
            from django.utils.dateparse import parse_date
            logs = logs.filter(timestamp__date__lte=parse_date(date_fin))
        except:
            pass

    # Limiter les résultats
    limit = int(request.GET.get('limit', 1000))
    logs = logs[:limit]

    # Créer un nouveau workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historique Audit"

    # Style de l'en-tête
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # En-têtes
    headers = [
        "Date/Heure", "Utilisateur", "Action", "Modèle",
        "ID Objet", "Objet", "Adresse IP", "User Agent"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Données
    for row_idx, log in enumerate(logs, 2):
        utilisateur = log.utilisateur.email if log.utilisateur else 'Système'

        ws.append([
            log.timestamp.strftime('%d/%m/%Y %H:%M:%S') if log.timestamp else '',
            utilisateur,
            log.get_action_display(),
            log.model_name,
            log.object_id,
            log.object_repr,
            log.ip_address or '',
            log.user_agent[:100] if log.user_agent else '',
        ])

    # Ajuster la largeur des colonnes
    column_widths = [20, 30, 20, 15, 25, 40, 15, 50]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    # Créer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Sauvegarder le workbook dans la réponse
    wb.save(response)

    return response
